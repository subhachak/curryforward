import os
import sys
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

os.environ["ADMIN_TOKEN"] = "test-token-123"
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

from fastapi.testclient import TestClient
from openpyxl import Workbook
from app.main import app
from app.db import SessionLocal, init_db
from app.seed_loader import load_seed_data

init_db()
_db = SessionLocal()
load_seed_data(_db)
_db.close()

client = TestClient(app)
ADMIN_HEADERS = {"X-Admin-Token": "test-token-123"}


CSV_TEXT = """name,category,cuisine_tags,servings,ingredients,steps,intro,tips
Sheet Chicken,dinner,"bengali, weeknight",4 servings,"200 g chicken; 1 tsp salt","Mix chicken; Cook until done",A sheet recipe,"Rest before serving"
,dessert,,2 servings,"1 cup flour","Bake it",Missing name row,
"""


def _fake_response(content: str):
    return SimpleNamespace(
        choices=[SimpleNamespace(message=SimpleNamespace(content=content))],
        usage={"prompt_tokens": 10, "completion_tokens": 5, "total_tokens": 15},
    )


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    first = workbook.active
    first.title = "Mains"
    first.append(["name", "category", "servings", "ingredients", "steps"])
    first.append(["Excel Chicken", "dinner", "4 servings", "200 g chicken; 1 tsp salt", "Mix; Cook"])
    second = workbook.create_sheet("Desserts")
    second.append(["title", "course", "yield", "ingredient_list", "directions"])
    second.append(["Excel Kheer", "dessert", "6 servings", "1 l milk; 50 g rice", "Simmer; Chill"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _recipe_sheet_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "American Buttercream"
    sheet.append(["", "Yelds", 1, 1, "Batch"])
    sheet.append([])
    sheet.append([1, "Vanilla", 1, 1, "Tablespoon", 15, 15, "ml"])
    sheet.append([2, "Confectioner's sugar", 1, 1, "lb", 445, 445, "Gram"])
    sheet.append([3, "Unsalted butter, room temperature", 1, 1, "Cup", 225, 225, "Gram"])
    sheet.append(["Instructions"])
    sheet.append(["Beat the butter until fluffy."])
    sheet.append(["Add sugar and beat until smooth."])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _freeform_recipe_sheet_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Honey Sriracha Wings"
    sheet.append(["25 chicken wings"])
    sheet.append([])
    sheet.append(["Ingredients"])
    sheet.append(["Chicken Wings"])
    sheet.append(["2.08 pound chicken wings patted dry"])
    sheet.append(["0.52 teaspoon garlic powder"])
    sheet.append(["Honey Sriracha Marinade"])
    sheet.append(["6.25 tablespoons honey"])
    sheet.append(["2.08 - 3.13 tablespoons sriracha"])
    sheet.append([])
    sheet.append(["Instructions"])
    sheet.append(["Preheat the oven to 450 F."])
    sheet.append(["Season the wings and bake until golden."])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _non_recipe_sheet_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cake Pan Conversion Chart"
    sheet.append([])
    sheet.append(["Recipe Calls For", "Volume", "Use Instead"])
    sheet.append(["1 (8-inch) round cake pan", "4 cups", "1 (9-inch) pie plate"])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_guest_cannot_preview_recipe_import():
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", CSV_TEXT, "text/csv")},
    )
    assert r.status_code == 403


def test_import_preview_parses_csv_rows(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", CSV_TEXT, "text/csv")},
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["valid_count"] == 1
    assert body["issue_count"] == 1
    assert body["source"] == "heuristic"
    assert body["file_type"] == "csv"
    first = body["rows"][0]
    assert first["name"] == "Sheet Chicken"
    assert first["base_servings_amount"] == 4
    assert first["cuisine_tags"] == ["bengali", "weeknight"]
    assert first["components"][0]["ingredients"][0]["name"] == "chicken"
    assert first["steps"][0]["instruction"] == "Mix chicken"


def test_import_preview_reads_excel_tabs(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={
            "file": (
                "recipes.xlsx",
                _xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["file_type"] == "xlsx"
    assert body["valid_count"] == 2
    assert [row["sheet_name"] for row in body["rows"]] == ["Mains", "Desserts"]
    assert body["rows"][0]["name"] == "Excel Chicken"
    assert body["rows"][1]["name"] == "Excel Kheer"


def test_import_preview_reads_recipe_per_sheet_workbook(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={
            "file": (
                "recipes.xlsx",
                _recipe_sheet_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["valid_count"] == 1
    assert body["issue_count"] == 0
    row = body["rows"][0]
    assert row["sheet_name"] == "American Buttercream"
    assert row["name"] == "American Buttercream"
    assert row["base_servings_amount"] == 1
    assert row["base_servings_unit"] == "Batch"
    assert row["components"][0]["ingredients"][0]["name"] == "Vanilla"
    assert row["components"][0]["ingredients"][0]["amount"] == 1
    assert row["components"][0]["ingredients"][0]["unit"] == "Tablespoon"
    assert row["steps"][0]["instruction"] == "Beat the butter until fluffy."


def test_import_preview_reads_freeform_recipe_sheet(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={
            "file": (
                "recipes.xlsx",
                _freeform_recipe_sheet_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["valid_count"] == 1
    row = body["rows"][0]
    assert row["name"] == "Honey Sriracha Wings"
    assert len(row["components"]) == 2
    assert row["components"][0]["component_name"] == "Chicken Wings"
    assert row["components"][0]["ingredients"][0]["name"] == "chicken wings patted dry"
    assert row["components"][1]["component_name"] == "Honey Sriracha Marinade"
    assert row["steps"][1]["instruction"] == "Season the wings and bake until golden."


def test_import_preview_skips_non_recipe_workbook_tabs(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={
            "file": (
                "recipes.xlsx",
                _non_recipe_sheet_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 400
    assert "Workbook does not contain any non-empty recipe rows" in r.json()["detail"]


def test_import_preview_uses_ai_mapper_when_available(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_litellm_configured", lambda: True)
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: True)
    monkeypatch.setattr(
        "app.routers.admin.litellm_completion",
        lambda **kwargs: _fake_response(
            """
            {
              "rows": [
                {
                  "sheet_name": null,
                  "row_number": 2,
                  "name": "AI Sheet Chicken",
                  "category": "dinner",
                  "cuisine_tags": ["bengali", "weeknight"],
                  "base_servings_amount": 4,
                  "base_servings_unit": "servings",
                  "intro": "AI mapped intro.",
                  "history": null,
                  "components": [{"component_name": "main", "ingredients": [{"name": "chicken", "amount": 200, "unit": "g"}]}],
                  "steps": [{"instruction": "Cook the chicken."}],
                  "tips": ["Rest before serving"],
                  "watch_outs": [],
                  "source_url": null
                },
                {
                  "sheet_name": null,
                  "row_number": 3,
                  "name": "",
                  "category": "dessert",
                  "cuisine_tags": [],
                  "base_servings_amount": 2,
                  "base_servings_unit": "servings",
                  "intro": null,
                  "history": null,
                  "components": [{"component_name": "main", "ingredients": [{"name": "flour", "amount": 1, "unit": "cup"}]}],
                  "steps": [{"instruction": "Bake it."}],
                  "tips": [],
                  "watch_outs": [],
                  "source_url": null
                }
              ]
            }
            """
        ),
    )
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", CSV_TEXT, "text/csv")},
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["source"] == "ai"
    assert body["rows"][0]["name"] == "AI Sheet Chicken"
    assert body["rows"][0]["intro"] == "AI mapped intro."
    assert body["rows"][1]["issues"] == ["Missing recipe name"]


def test_import_preview_falls_back_when_ai_mapper_fails(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_litellm_configured", lambda: True)
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: True)

    def fail_completion(**kwargs):
        raise RuntimeError("provider timed out")

    monkeypatch.setattr("app.routers.admin.litellm_completion", fail_completion)
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", CSV_TEXT, "text/csv")},
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["source"] == "heuristic"
    assert "provider timed out" in body["ai_error"]
    assert body["valid_count"] == 1


def test_import_preview_skips_ai_for_large_uploads(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_litellm_configured", lambda: True)
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: True)

    def fail_if_called(**kwargs):
        raise AssertionError("AI mapper should not be called for large imports")

    monkeypatch.setattr("app.routers.admin.litellm_completion", fail_if_called)
    rows = [
        "name,category,servings,ingredients,steps",
        *[
            f"Recipe {idx},dinner,4 servings,\"1 cup rice; 1 tsp salt\",\"Rinse rice; Cook rice\""
            for idx in range(30)
        ],
    ]
    r = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", "\n".join(rows), "text/csv")},
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["source"] == "heuristic"
    assert body["valid_count"] == 30
    assert "AI mapping skipped" in body["ai_error"]


def test_import_commit_creates_draft_for_valid_rows_only(monkeypatch):
    monkeypatch.setattr("app.routers.admin.is_model_available", lambda model: False)
    preview = client.post(
        "/api/admin/recipes/import/preview",
        files={"file": ("recipes.csv", CSV_TEXT, "text/csv")},
        headers=ADMIN_HEADERS,
    ).json()
    r = client.post(
        "/api/admin/recipes/import/commit",
        json={"rows": preview["rows"]},
        headers=ADMIN_HEADERS,
    )
    assert r.status_code == 200
    body = r.json()
    assert len(body["created"]) == 1
    assert len(body["skipped"]) == 1
    recipe_id = body["created"][0]["recipe_id"]
    draft = client.get(f"/api/recipes/research/{recipe_id}", headers=ADMIN_HEADERS).json()
    assert draft["status"] == "draft"
    assert draft["source"] == "imported"
    assert draft["name"] == "Sheet Chicken"
