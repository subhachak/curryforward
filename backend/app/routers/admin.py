"""
Admin dashboard surface — the unified published+draft recipe list (with
analytics), Trash (soft-deleted recipes), restore, and permanent purge.

Distinct from routers/recipes.py (guest-facing CRUD/chat) and
routers/research.py (the research workflow specifically): this router is
purely about managing recipes that already exist, regardless of how they
were created (manual, research, or fork).
"""
from __future__ import annotations

import csv
import io
import json
import re
import uuid

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..auth import require_admin
from ..db import get_db
from ..llm_client import is_litellm_configured, is_model_available, litellm_completion
from ..models import AdminAuditLog, LLMUsageLog, RecipeAnalytics, RecipeFeedback, RecipeVersion
from ..nutrition import compute_nutrition
from ..services.audit import audit_admin_action
from ..services.ingredient_canonical import normalize_components_to_grams
from ..services.llm_settings import available_models, get_llm_settings, resolve_task_model, set_llm_setting
from ..services.llm_usage import record_llm_usage
from ..services.recipe_versions import fork_recipe_version

router = APIRouter(prefix="/api/admin")

MAX_AI_IMPORT_ROWS = 25
MAX_AI_IMPORT_CELL_CHARS = 1200
AI_IMPORT_TIMEOUT_SECONDS = 20


class FeedbackDecision(BaseModel):
    approved: bool


class LLMSettingUpdate(BaseModel):
    model: str


class CopyRewriteRequest(BaseModel):
    field_label: str
    text: str
    instruction: str | None = None
    recipe_context: str | None = None


class CopyRewriteResponse(BaseModel):
    text: str


class RecipeImportRow(BaseModel):
    sheet_name: str | None = None
    row_number: int
    name: str
    category: str | None = None
    cuisine_tags: list[str] = Field(default_factory=list)
    base_servings_amount: float | None = None
    base_servings_unit: str = "servings"
    intro: str | None = None
    history: str | None = None
    components: list[dict] = Field(default_factory=list)
    steps: list[dict] = Field(default_factory=list)
    tips: list[str] = Field(default_factory=list)
    watch_outs: list[str] = Field(default_factory=list)
    source_url: str | None = None
    issues: list[str] = Field(default_factory=list)


class RecipeImportCommitRequest(BaseModel):
    rows: list[RecipeImportRow]


def _rewrite_copy_text(req: CopyRewriteRequest, model: str) -> tuple[str, object]:
    field_label = req.field_label.strip()[:120] or "field"
    source_text = req.text.strip()
    instruction = (req.instruction or "").strip() or "Rewrite this into polished, user-friendly copy."
    context = (req.recipe_context or "").strip()[:600]
    context_line = f"Nearby context: {context}\n" if context else ""
    if not source_text:
        raise HTTPException(400, "Text is required")
    response = litellm_completion(
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a precise recipe copy editor inside CurryForward. Rewrite only "
                    "the requested field into clear, warm, publishable recipe copy. Keep the "
                    "same factual meaning. Do not invent facts, ingredients, timings, dietary "
                    "claims, history, or provenance. Return only the rewritten field text, no "
                    "quotes, markdown, labels, or explanation. Preserve list-like formatting "
                    "when the input is a newline-separated list."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Field: {field_label}\n"
                    f"Admin direction: {instruction}\n"
                    f"{context_line}"
                    f"Current text:\n{source_text}"
                ),
            },
        ],
        temperature=0.4,
    )
    return (response.choices[0].message.content or "").strip(), response


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _first_value(row: dict[str, str], *names: str) -> str:
    normalized = {key.strip().lower().replace(" ", "_"): value for key, value in row.items()}
    for name in names:
        value = normalized.get(name)
        if value is not None and value.strip():
            return value.strip()
    return ""


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


IMPORT_HEADER_ALIASES = {
    "name",
    "title",
    "recipe",
    "recipe_name",
    "category",
    "course",
    "cuisine_tags",
    "tags",
    "cuisine",
    "servings",
    "yield",
    "base_servings",
    "ingredients",
    "ingredient_list",
    "steps",
    "instructions",
    "method",
    "directions",
    "intro",
    "description",
    "summary",
    "history",
    "notes",
    "tips",
    "tips_and_tricks",
    "watch_outs",
    "watchouts",
    "warnings",
    "source_url",
    "url",
    "source",
}

SECTION_ALIASES = {
    "name": "name",
    "title": "name",
    "recipe": "name",
    "recipe_name": "name",
    "category": "category",
    "course": "category",
    "yelds": "servings",
    "yields": "servings",
    "serves": "servings",
    "servings": "servings",
    "yield": "servings",
    "base_servings": "servings",
    "ingredients": "ingredients",
    "ingredient_list": "ingredients",
    "steps": "steps",
    "instruction": "steps",
    "instructions": "steps",
    "method": "steps",
    "directions": "steps",
    "intro": "intro",
    "description": "intro",
    "summary": "intro",
    "history": "history",
    "notes": "history",
    "tips": "tips",
    "tips_and_tricks": "tips",
    "watch_outs": "watch_outs",
    "watchouts": "watch_outs",
    "warnings": "watch_outs",
    "source_url": "source_url",
    "url": "source_url",
    "source": "source_url",
}


def _parse_list(value: str, *, split_commas: bool = False) -> list[str]:
    if not value:
        return []
    pattern = r"[\n;]+|,\s*(?=[^\d])" if split_commas else r"[\n;]+"
    parts = re.split(pattern, value)
    return [part.strip(" -\t") for part in parts if part.strip(" -\t")]


def _parse_servings(value: str) -> tuple[float | None, str]:
    if not value:
        return None, "servings"
    match = re.match(r"^\s*([\d.]+)\s*(.*)$", value)
    if not match:
        return None, value.strip() or "servings"
    return float(match.group(1)), (match.group(2).strip() or "servings")


def _parse_optional_float(value: str) -> float | None:
    value = value.strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _is_index_cell(value: str) -> bool:
    number = _parse_optional_float(value)
    return number is not None and number > 0 and number.is_integer()


def _parse_ingredient_line(line: str) -> dict:
    text = line.strip(" -\t")
    match = re.match(r"^([\d./\s]+)?\s*([A-Za-z]+|cups?|tbsp|tsp|teaspoons?|tablespoons?)?\s+(.+)$", text)
    amount = None
    unit = ""
    name = text
    if match and match.group(3):
        raw_amount = (match.group(1) or "").strip()
        if raw_amount:
            try:
                if "/" in raw_amount:
                    pieces = raw_amount.split()
                    total = 0.0
                    for piece in pieces:
                        if "/" in piece:
                            numerator, denominator = piece.split("/", 1)
                            total += float(numerator) / float(denominator)
                        else:
                            total += float(piece)
                    amount = total
                else:
                    amount = float(raw_amount)
            except ValueError:
                amount = None
        unit = (match.group(2) or "").strip()
        name = match.group(3).strip()
    return {"name": name, "amount": amount, "unit": unit}


def _parse_components(value: str) -> list[dict]:
    ingredients = [_parse_ingredient_line(line) for line in _parse_list(value)]
    ingredients = [ingredient for ingredient in ingredients if ingredient["name"]]
    return [{"component_name": "main", "ingredients": ingredients}] if ingredients else []


def _parse_steps(value: str) -> list[dict]:
    return [{"instruction": step} for step in _parse_list(value)]


def _import_issues(row: RecipeImportRow) -> list[str]:
    issues = []
    if not row.name.strip():
        issues.append("Missing recipe name")
    if not row.components or not any((component.get("ingredients") or []) for component in row.components):
        issues.append("Missing ingredients")
    if not row.steps:
        issues.append("Missing steps")
    return issues


def _normalize_import_row(row: dict[str, str], row_number: int, sheet_name: str | None = None) -> RecipeImportRow:
    name = _first_value(row, "name", "title", "recipe", "recipe_name")
    servings_amount, servings_unit = _parse_servings(_first_value(row, "servings", "yield", "base_servings"))
    ingredients = _first_value(row, "ingredients", "ingredient_list")
    steps = _first_value(row, "steps", "instructions", "method", "directions")
    normalized = RecipeImportRow(
        sheet_name=sheet_name,
        row_number=row_number,
        name=name or f"Imported recipe row {row_number}",
        category=_first_value(row, "category", "course") or None,
        cuisine_tags=_parse_list(_first_value(row, "cuisine_tags", "tags", "cuisine"), split_commas=True),
        base_servings_amount=servings_amount,
        base_servings_unit=servings_unit,
        intro=_first_value(row, "intro", "description", "summary") or None,
        history=_first_value(row, "history", "notes") or None,
        components=_parse_components(ingredients),
        steps=_parse_steps(steps),
        tips=_parse_list(_first_value(row, "tips", "tips_and_tricks")),
        watch_outs=_parse_list(_first_value(row, "watch_outs", "watchouts", "warnings")),
        source_url=_first_value(row, "source_url", "url", "source") or None,
    )
    if not name:
        normalized.issues.append("Missing recipe name")
    if not ingredients:
        normalized.issues.append("Missing ingredients")
    if not steps:
        normalized.issues.append("Missing steps")
    return normalized


def _looks_like_table_header(values: tuple | list) -> bool:
    normalized_headers = {_normalize_header(_cell_to_text(value)) for value in values if _cell_to_text(value)}
    return len(normalized_headers & IMPORT_HEADER_ALIASES) >= 2


def _read_structured_recipe_sheet(sheet) -> tuple[dict[str, str] | None, RecipeImportRow | None]:
    components: list[dict] = []
    current_component = {"component_name": "main", "ingredients": []}
    steps: list[dict] = []
    servings = ""
    in_ingredients = False
    in_steps = False
    saw_structured_row = False
    raw_cells = []

    def flush_component():
        nonlocal current_component
        if current_component["ingredients"]:
            components.append(current_component)
        current_component = {"component_name": "main", "ingredients": []}

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [_cell_to_text(value) for value in values]
        non_empty = [cell for cell in cells if cell]
        if not non_empty:
            continue
        raw_cells.append(" | ".join(non_empty))
        normalized = [_normalize_header(cell) for cell in cells]

        if any(label in {"yelds", "yields", "serves", "servings"} for label in normalized):
            label_index = next(
                idx for idx, label in enumerate(normalized) if label in {"yelds", "yields", "serves", "servings"}
            )
            tail = [cell for cell in cells[label_index + 1 :] if cell]
            numeric_cells = [cell for cell in tail if _parse_optional_float(cell) is not None]
            unit = next((cell for cell in tail if _parse_optional_float(cell) is None), "servings")
            amount = numeric_cells[-1] if numeric_cells else ""
            servings = f"{amount} {unit}".strip()
            saw_structured_row = True
            continue

        if any(label in {"steps", "instruction", "instructions", "method", "directions"} for label in normalized):
            flush_component()
            in_ingredients = False
            in_steps = True
            saw_structured_row = True
            continue

        if any(label in {"ingredients", "ingredient_list"} for label in normalized):
            in_ingredients = True
            in_steps = False
            saw_structured_row = True
            continue

        if in_steps:
            if len(non_empty) == 1 and not re.search(r"[.!?]$", non_empty[0]):
                continue
            instruction = cells[1] if len(cells) > 1 and cells[1] else non_empty[-1]
            if instruction and _normalize_header(instruction) not in SECTION_ALIASES:
                steps.append({"instruction": instruction})
                saw_structured_row = True
            continue

        if cells and _is_index_cell(cells[0]) and len(cells) > 1 and cells[1]:
            if len(non_empty) <= 2 and (len(cells[1]) > 80 or re.search(r"[.!?]$", cells[1])):
                flush_component()
                in_ingredients = False
                in_steps = True
                steps.append({"instruction": cells[1]})
                saw_structured_row = True
                continue
            amount_text = cells[3] if len(cells) > 3 and cells[3] else (cells[2] if len(cells) > 2 else "")
            unit = cells[4] if len(cells) > 4 else ""
            ingredient = {
                "name": cells[1],
                "amount": _parse_optional_float(amount_text),
                "unit": unit,
            }
            unit_options = []
            if len(cells) > 2 and cells[2] and cells[2] != amount_text:
                unit_options.append({"amount": _parse_optional_float(cells[2]), "unit": unit})
            if len(cells) > 7 and cells[6] and cells[7]:
                unit_options.append({"amount": _parse_optional_float(cells[6]), "unit": cells[7]})
            if unit_options:
                ingredient["unit_options"] = unit_options
            current_component["ingredients"].append(ingredient)
            saw_structured_row = True
            continue

        if in_ingredients:
            text = " ".join(non_empty).strip()
            parsed = _parse_ingredient_line(text)
            if parsed["amount"] is None and (not parsed["unit"] or not re.search(r"\d", text)):
                flush_component()
                current_component = {"component_name": text, "ingredients": []}
            else:
                current_component["ingredients"].append(parsed)
            saw_structured_row = True
            continue

        if len(non_empty) == 1 and not _parse_optional_float(non_empty[0]) and not re.search(r"\d", non_empty[0]):
            label = non_empty[0].strip()
            normalized_label = _normalize_header(label)
            if normalized_label not in SECTION_ALIASES and not label.lower().startswith("http"):
                flush_component()
                current_component = {"component_name": label, "ingredients": []}

    flush_component()
    if not saw_structured_row or (not components and not steps):
        return None, None

    row = {
        "name": sheet.title,
        "servings": servings,
        "ingredients": "",
        "steps": "",
    }
    raw = {"_sheet_name": sheet.title, "_layout": "structured_recipe", "_raw_text": "\n".join(raw_cells), **row}
    normalized = RecipeImportRow(
        sheet_name=sheet.title,
        row_number=1,
        name=sheet.title,
        category=None,
        cuisine_tags=[],
        base_servings_amount=_parse_servings(servings)[0],
        base_servings_unit=_parse_servings(servings)[1],
        components=components,
        steps=steps,
        tips=[],
        watch_outs=[],
        source_url=None,
    )
    normalized.issues = _import_issues(normalized)
    return raw, normalized


def _append_block_value(blocks: dict[str, list[str]], key: str, value: str):
    value = value.strip()
    if value:
        blocks.setdefault(key, []).append(value)


def _read_sheet_as_recipe(sheet) -> tuple[dict[str, str] | None, RecipeImportRow | None]:
    structured_raw, structured_row = _read_structured_recipe_sheet(sheet)
    if structured_raw and structured_row:
        return structured_raw, structured_row

    raw_cells = []
    scalar_values: dict[str, str] = {}
    blocks: dict[str, list[str]] = {}
    current_section: str | None = None
    first_title: str | None = None
    first_row_number = 1

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [_cell_to_text(value) for value in values]
        non_empty = [cell for cell in cells if cell]
        if not non_empty:
            continue
        if first_title is None:
            first_row_number = row_number
            first_title = non_empty[0]
        raw_cells.append(" | ".join(non_empty))

        first = non_empty[0]
        inline_label = re.match(r"^([^:]{1,60}):\s*(.+)$", first)
        if inline_label:
            label = _normalize_header(inline_label.group(1))
            canonical = SECTION_ALIASES.get(label)
            if canonical:
                _append_block_value(blocks, canonical, inline_label.group(2))
                current_section = canonical if canonical in {"ingredients", "steps", "tips", "watch_outs"} else None
                continue

        label = _normalize_header(first)
        canonical = SECTION_ALIASES.get(label)
        rest = " ".join(non_empty[1:]).strip()
        if canonical and rest:
            scalar_values[canonical] = rest if canonical not in blocks else scalar_values.get(canonical, rest)
            _append_block_value(blocks, canonical, rest)
            current_section = canonical if canonical in {"ingredients", "steps", "tips", "watch_outs"} else None
            continue
        if canonical and len(non_empty) == 1:
            current_section = canonical
            continue
        if current_section:
            _append_block_value(blocks, current_section, " ".join(non_empty))

    if not raw_cells:
        return None, None
    if not blocks.get("ingredients") and not blocks.get("steps"):
        return None, None

    row = {
        "name": scalar_values.get("name") or sheet.title,
        "category": scalar_values.get("category", ""),
        "servings": scalar_values.get("servings", ""),
        "ingredients": "\n".join(blocks.get("ingredients", [])),
        "steps": "\n".join(blocks.get("steps", [])),
        "intro": scalar_values.get("intro") or "\n".join(blocks.get("intro", [])),
        "history": scalar_values.get("history") or "\n".join(blocks.get("history", [])),
        "tips": "\n".join(blocks.get("tips", [])),
        "watch_outs": "\n".join(blocks.get("watch_outs", [])),
        "source_url": scalar_values.get("source_url", ""),
    }
    raw = {"_sheet_name": sheet.title, "_layout": "sheet_recipe", "_raw_text": "\n".join(raw_cells), **row}
    return raw, _normalize_import_row(row, first_row_number, sheet.title)


def _read_csv_upload(raw: bytes) -> tuple[list[dict[str, str]], list[RecipeImportRow]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(400, "CSV needs a header row")
    raw_rows = [dict(row) for row in reader]
    return raw_rows, [_normalize_import_row(row, idx) for idx, row in enumerate(raw_rows, start=2)]


def _read_xlsx_upload(raw: bytes) -> tuple[list[dict[str, str]], list[RecipeImportRow]]:
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel workbook: {e}")
    raw_rows = []
    normalized_rows = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (idx for idx, values in enumerate(rows) if _looks_like_table_header(values)),
            None,
        )
        if header_index is None:
            raw, row = _read_sheet_as_recipe(sheet)
            if raw and row:
                raw_rows.append(raw)
                normalized_rows.append(row)
            continue
        headers = [_cell_to_text(value) or f"column_{idx + 1}" for idx, value in enumerate(rows[header_index])]
        for offset, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(_cell_to_text(value) for value in values):
                continue
            row = {
                headers[idx]: _cell_to_text(values[idx]) if idx < len(values) else ""
                for idx in range(len(headers))
            }
            raw_rows.append({"_sheet_name": sheet.title, **row})
            normalized_rows.append(_normalize_import_row(row, offset, sheet.title))
    if not raw_rows:
        raise HTTPException(400, "Workbook does not contain any non-empty recipe rows")
    return raw_rows, normalized_rows


async def _read_spreadsheet_upload(file: UploadFile) -> tuple[list[dict[str, str]], list[RecipeImportRow], str]:
    raw = await file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        raw_rows, rows = _read_xlsx_upload(raw)
        return raw_rows, rows, "xlsx"
    if filename.endswith(".csv") or file.content_type in {"text/csv", "application/csv"}:
        raw_rows, rows = _read_csv_upload(raw)
        return raw_rows, rows, "csv"
    raise HTTPException(400, "Upload a .xlsx workbook or .csv file")


def _extract_json_payload(text: str) -> dict:
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if not match:
            raise
        payload = json.loads(match.group(0))
    if isinstance(payload, list):
        return {"rows": payload}
    if not isinstance(payload, dict):
        raise ValueError("Import mapper returned non-object JSON")
    return payload


def _coerce_import_row(candidate: dict, fallback: RecipeImportRow) -> RecipeImportRow:
    merged = fallback.model_dump()
    for key in [
        "name",
        "category",
        "cuisine_tags",
        "base_servings_amount",
        "base_servings_unit",
        "intro",
        "history",
        "components",
        "steps",
        "tips",
        "watch_outs",
        "source_url",
    ]:
        if key in candidate and candidate[key] is not None:
            merged[key] = candidate[key]
    merged["row_number"] = fallback.row_number
    merged["sheet_name"] = fallback.sheet_name
    row = RecipeImportRow.model_validate(merged)
    row.issues = _import_issues(row)
    return row


def _ai_map_import_rows(
    raw_rows: list[dict[str, str]],
    fallback_rows: list[RecipeImportRow],
    model: str,
    role: str,
) -> tuple[list[RecipeImportRow], object]:
    sample = [
        {
            "sheet_name": fallback.sheet_name,
            "row_number": fallback.row_number,
            "cells": {key: value[:MAX_AI_IMPORT_CELL_CHARS] for key, value in raw.items()},
        }
        for fallback, raw in zip(fallback_rows, raw_rows)
    ]
    response = litellm_completion(
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "You map spreadsheet recipe rows into CurryForward draft recipe JSON. "
                    "Return ONLY valid JSON with one top-level key: rows. rows must be an "
                    "array with one object per input row, in the same order. Preserve the "
                    "given sheet_name and row_number. Use this schema for each object: "
                    "sheet_name, row_number, name, category, cuisine_tags, base_servings_amount, base_servings_unit, "
                    "intro, history, components, steps, tips, watch_outs, source_url. "
                    "components is an array of {component_name, ingredients}; each ingredient "
                    "is {name, amount, unit}. steps is an array of {instruction}. tips, "
                    "watch_outs, and cuisine_tags are string arrays. If a field is absent, "
                    "use null or an empty array. Do not invent facts that are not supported "
                    "by the cells."
                ),
            },
            {
                "role": "user",
                "content": json.dumps({"rows": sample}, ensure_ascii=False),
            },
        ],
        temperature=0.1,
        timeout=AI_IMPORT_TIMEOUT_SECONDS,
    )
    payload = _extract_json_payload(response.choices[0].message.content or "{}")
    mapped = payload.get("rows") or []
    if not isinstance(mapped, list):
        raise ValueError("Import mapper returned rows in an invalid format")
    by_location = {
        (str(item.get("sheet_name") or ""), int(item.get("row_number"))): item
        for item in mapped
        if isinstance(item, dict) and str(item.get("row_number", "")).isdigit()
    }
    rows = [
        _coerce_import_row(by_location.get((str(row.sheet_name or ""), row.row_number), {}), row)
        for row in fallback_rows
    ]
    record_llm_usage(task="recipe_import", model=model, role=role, response=response)
    return rows, response


@router.post("/recipes/import/preview")
async def preview_recipe_import(
    file: UploadFile = File(...),
    model: str | None = Form(default=None),
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    raw_rows, rows, file_type = await _read_spreadsheet_upload(file)
    source = "heuristic"
    ai_error = None
    resolved_model = resolve_task_model("recipe_import", db, model)
    if len(rows) > MAX_AI_IMPORT_ROWS:
        ai_error = (
            f"AI mapping skipped for {len(rows)} rows to keep preview responsive. "
            f"Upload {MAX_AI_IMPORT_ROWS} rows or fewer at a time for AI-assisted mapping."
        )
    elif is_litellm_configured() and is_model_available(resolved_model):
        try:
            rows, _response = _ai_map_import_rows(raw_rows, rows, resolved_model, role)
            source = "ai"
        except Exception as e:
            ai_error = str(e)
            record_llm_usage(task="recipe_import", model=resolved_model, role=role, status="error", error=ai_error)
    return {
        "rows": [row.model_dump() for row in rows],
        "valid_count": sum(1 for row in rows if not row.issues),
        "issue_count": sum(1 for row in rows if row.issues),
        "source": source,
        "model": resolved_model,
        "ai_error": ai_error,
        "file_type": file_type,
    }


@router.post("/recipes/import/commit")
def commit_recipe_import(
    req: RecipeImportCommitRequest,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    created = []
    skipped = []
    for row in req.rows:
        if row.issues:
            skipped.append({
                "sheet_name": row.sheet_name,
                "row_number": row.row_number,
                "name": row.name,
                "issues": row.issues,
            })
            continue
        components = normalize_components_to_grams(row.components)
        recipe = RecipeVersion(
            recipe_id=f"import-{uuid.uuid4().hex[:8]}",
            parent_version_id=None,
            lineage="imported",
            name=row.name,
            category=row.category,
            cuisine_tags=row.cuisine_tags,
            base_servings_amount=row.base_servings_amount,
            base_servings_unit=row.base_servings_unit,
            components=components,
            steps=row.steps,
            nutrition=compute_nutrition(components, db),
            intro=row.intro,
            history=row.history,
            tips=row.tips,
            watch_outs=row.watch_outs,
            status="draft",
            source="imported",
            is_current_head=True,
            notes=(
                f"Imported from spreadsheet"
                f"{f' sheet {row.sheet_name}' if row.sheet_name else ''}"
                f" row {row.row_number}"
                + (f"\nSource: {row.source_url}" if row.source_url else "")
            ),
            research_conversation={"messages": [], "pending_tool_use": None},
        )
        db.add(recipe)
        created.append({
            "recipe_id": recipe.recipe_id,
            "name": recipe.name,
            "sheet_name": row.sheet_name,
            "row_number": row.row_number,
        })
    audit_admin_action(
        db,
        action="recipes_imported",
        target_type="recipe",
        request=request,
        details={"created": len(created), "skipped": len(skipped)},
    )
    db.commit()
    return {"created": created, "skipped": skipped}


@router.get("/recipes")
def list_all_recipes(db: Session = Depends(get_db), role: str = Depends(require_admin)):
    """Unified dashboard list: every non-deleted recipe (published + draft),
    with status and analytics attached."""
    rows = (
        db.query(RecipeVersion)
        .filter(RecipeVersion.is_current_head == True, RecipeVersion.deleted_at.is_(None))  # noqa: E712
        .order_by(RecipeVersion.updated_at.desc())
        .all()
    )
    recipe_ids = [r.recipe_id for r in rows]
    analytics = {
        a.recipe_id: a
        for a in db.query(RecipeAnalytics).filter(RecipeAnalytics.recipe_id.in_(recipe_ids)).all()
    }
    result = []
    for r in rows:
        first_published_at = (
            db.query(func.min(RecipeVersion.created_at))
            .filter(RecipeVersion.recipe_id == r.recipe_id, RecipeVersion.status == "published")
            .scalar()
        )
        result.append({
            "recipe_id": r.recipe_id,
            "version_id": r.version_id,
            "name": r.name,
            "category": r.category,
            "status": r.status or "published",
            "lineage": r.lineage,
            "hero_image_url": r.hero_image_url,
            "intro": r.intro,
            "first_published_at": first_published_at.isoformat() if first_published_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            "view_count": analytics[r.recipe_id].view_count if r.recipe_id in analytics else 0,
            "download_count": analytics[r.recipe_id].download_count if r.recipe_id in analytics else 0,
            "like_count": analytics[r.recipe_id].like_count if r.recipe_id in analytics else 0,
        })
    return result


@router.get("/llm-settings")
def list_llm_settings(db: Session = Depends(get_db), role: str = Depends(require_admin)):
    return {
        "settings": get_llm_settings(db),
        "models": available_models(),
    }


@router.put("/llm-settings/{key}")
def update_llm_setting(
    key: str,
    req: LLMSettingUpdate,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    try:
        result = set_llm_setting(db, key, req.model)
        audit_admin_action(
            db,
            action="llm_setting_updated",
            target_type="llm_setting",
            target_id=key,
            request=request,
            details={"model": req.model},
        )
        db.commit()
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/rewrite", response_model=CopyRewriteResponse)
def rewrite_admin_copy(
    req: CopyRewriteRequest,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    if not is_litellm_configured():
        raise HTTPException(400, "litellm is not installed — check backend/requirements.txt")
    model = resolve_task_model("copy_rewrite", db)
    if not is_model_available(model):
        raise HTTPException(400, f"No API key configured for model '{model}' — add it to backend/.env")
    try:
        text, response = _rewrite_copy_text(req, model)
    except HTTPException:
        raise
    except Exception as e:
        record_llm_usage(task="copy_rewrite", model=model, role=role, status="error", error=str(e))
        raise HTTPException(500, f"Rewrite failed: {e}")
    record_llm_usage(task="copy_rewrite", model=model, role=role, response=response)
    if not text:
        raise HTTPException(500, "Rewrite returned an empty response")
    audit_admin_action(
        db,
        action="copy_rewrite_generated",
        target_type="admin",
        request=request,
        details={"field": req.field_label.strip()[:120]},
    )
    db.commit()
    return {"text": text}


@router.post("/recipes/{recipe_id}/edit-draft")
def create_edit_draft(
    recipe_id: str,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    """Dashboard-only edit entry point.

    Draft recipes can be edited in place. Published recipes get a linked draft
    working copy; publishing that copy later can either replace the original or
    keep both recipes.
    """
    current = (
        db.query(RecipeVersion)
        .filter(
            RecipeVersion.recipe_id == recipe_id,
            RecipeVersion.is_current_head == True,  # noqa: E712
            RecipeVersion.deleted_at.is_(None),
        )
        .first()
    )
    if not current:
        raise HTTPException(404, "Recipe not found")
    if (current.status or "published") == "draft":
        audit_admin_action(
            db,
            action="draft_opened",
            target_type="recipe",
            target_id=recipe_id,
            request=request,
            details={"created": False},
        )
        db.commit()
        return {
            "draft": current.to_research_dict(),
            "created": False,
            "note": "This draft is editable in place.",
        }

    existing = (
        db.query(RecipeVersion)
        .filter(
            RecipeVersion.parent_version_id == current.version_id,
            RecipeVersion.status == "draft",
            RecipeVersion.source == "revision_draft",
            RecipeVersion.is_current_head == True,  # noqa: E712
            RecipeVersion.deleted_at.is_(None),
        )
        .order_by(RecipeVersion.updated_at.desc())
        .first()
    )
    if existing:
        audit_admin_action(
            db,
            action="edit_draft_opened",
            target_type="recipe",
            target_id=recipe_id,
            request=request,
            details={"draft_recipe_id": existing.recipe_id, "created": False},
        )
        db.commit()
        return {
            "draft": existing.to_research_dict(),
            "created": False,
            "note": "Opened an existing draft copy of the published recipe.",
        }

    draft = fork_recipe_version(current)
    draft.name = f"{current.name} (draft edit)"
    draft.source = "revision_draft"
    db.add(draft)
    audit_admin_action(
        db,
        action="edit_draft_created",
        target_type="recipe",
        target_id=recipe_id,
        request=request,
        details={"draft_recipe_id": draft.recipe_id},
    )
    db.commit()
    return {
        "draft": draft.to_research_dict(),
        "created": True,
        "note": "Created a draft copy. The published recipe remains live until you choose how to publish this draft.",
    }


@router.get("/recipes/trash")
def list_trash(db: Session = Depends(get_db), role: str = Depends(require_admin)):
    """Soft-deleted recipes, most recently deleted first."""
    rows = (
        db.query(RecipeVersion)
        .filter(RecipeVersion.is_current_head == True, RecipeVersion.deleted_at.isnot(None))  # noqa: E712
        .order_by(RecipeVersion.deleted_at.desc())
        .all()
    )
    return [
        {
            "recipe_id": r.recipe_id,
            "version_id": r.version_id,
            "name": r.name,
            "category": r.category,
            "deleted_at": r.deleted_at.isoformat() if r.deleted_at else None,
        }
        for r in rows
    ]


@router.get("/feedback/pending")
def list_pending_feedback(db: Session = Depends(get_db), role: str = Depends(require_admin)):
    rows = (
        db.query(RecipeFeedback, RecipeVersion.name)
        .join(
            RecipeVersion,
            (RecipeVersion.recipe_id == RecipeFeedback.recipe_id)
            & (RecipeVersion.is_current_head == True),  # noqa: E712
        )
        .filter(RecipeFeedback.status == "pending_review")
        .order_by(RecipeFeedback.created_at.desc())
        .all()
    )
    return [
        {
            **feedback.to_dict(),
            "recipe_name": recipe_name,
        }
        for feedback, recipe_name in rows
    ]


@router.post("/feedback/{feedback_id}/decide")
def decide_feedback(
    feedback_id: str,
    req: FeedbackDecision,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    row = db.query(RecipeFeedback).filter(RecipeFeedback.feedback_id == feedback_id).first()
    if not row:
        raise HTTPException(404, "Feedback not found")
    row.status = "approved" if req.approved else "rejected"
    audit_admin_action(
        db,
        action="feedback_decided",
        target_type="feedback",
        target_id=feedback_id,
        request=request,
        details={"approved": req.approved, "recipe_id": row.recipe_id},
    )
    db.commit()
    db.refresh(row)
    return row.to_dict()


@router.post("/recipes/{recipe_id}/restore")
def restore_recipe(
    recipe_id: str,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    """Un-deletes a trashed recipe — clears deleted_at, no other state
    changes. It comes back as whatever status it had when trashed (always
    "draft", since delete is gated to drafts-only)."""
    row = (
        db.query(RecipeVersion)
        .filter(RecipeVersion.recipe_id == recipe_id, RecipeVersion.is_current_head == True)  # noqa: E712
        .first()
    )
    if not row:
        raise HTTPException(404, "Recipe not found")
    if row.deleted_at is None:
        raise HTTPException(400, "This recipe is not in the trash")
    row.deleted_at = None
    audit_admin_action(db, action="recipe_restored", target_type="recipe", target_id=recipe_id, request=request)
    db.commit()
    return row.to_dict()


@router.delete("/recipes/{recipe_id}/purge")
def purge_recipe(
    recipe_id: str,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Depends(require_admin),
):
    """Permanent hard-delete — removes every version row for this recipe_id,
    plus its analytics row. Only allowed once the current-head row is
    already soft-deleted (in Trash), so this can't be used to bypass the
    publish-then-unpublish-then-delete gate on a live recipe."""
    current = (
        db.query(RecipeVersion)
        .filter(RecipeVersion.recipe_id == recipe_id, RecipeVersion.is_current_head == True)  # noqa: E712
        .first()
    )
    if not current:
        raise HTTPException(404, "Recipe not found")
    if current.deleted_at is None:
        raise HTTPException(400, "Only trashed recipes can be permanently purged.")
    versions = db.query(RecipeVersion).filter(RecipeVersion.recipe_id == recipe_id).all()
    for v in versions:
        db.delete(v)
    db.query(RecipeAnalytics).filter(RecipeAnalytics.recipe_id == recipe_id).delete()
    db.query(RecipeFeedback).filter(RecipeFeedback.recipe_id == recipe_id).delete()
    audit_admin_action(db, action="recipe_purged", target_type="recipe", target_id=recipe_id, request=request)
    db.commit()
    return {"purged": recipe_id}


@router.get("/audit-log")
def list_audit_log(limit: int = 100, db: Session = Depends(get_db), role: str = Depends(require_admin)):
    limit = max(1, min(limit, 500))
    rows = db.query(AdminAuditLog).order_by(AdminAuditLog.created_at.desc()).limit(limit).all()
    return [row.to_dict() for row in rows]


@router.get("/llm-usage")
def list_llm_usage(limit: int = 100, db: Session = Depends(get_db), role: str = Depends(require_admin)):
    limit = max(1, min(limit, 500))
    rows = db.query(LLMUsageLog).order_by(LLMUsageLog.created_at.desc()).limit(limit).all()
    totals = (
        db.query(
            LLMUsageLog.task,
            LLMUsageLog.model,
            func.count(LLMUsageLog.usage_id),
            func.sum(LLMUsageLog.total_tokens),
        )
        .group_by(LLMUsageLog.task, LLMUsageLog.model)
        .all()
    )
    return {
        "items": [row.to_dict() for row in rows],
        "summary": [
            {
                "task": task,
                "model": model,
                "call_count": count,
                "total_tokens": int(total_tokens or 0),
            }
            for task, model, count, total_tokens in totals
        ],
    }
